import ast
import os
import sys
import math
import time
import traceback
from collections import defaultdict
from itertools import combinations
from pysat.formula import CNF
from pysat.solvers import Glucose3
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from zipfile import BadZipFile
from threading import Thread, Event


def write_to_excel(result_dict, folder):
    df = pd.DataFrame([result_dict])
    date = datetime.now().strftime('%Y-%m-%d')
    output_dir = f'out/{folder}'
    if not os.path.exists(output_dir): os.makedirs(output_dir)
    file_path = f'{output_dir}/results_{date}_2025.xlsx'

    if os.path.exists(file_path):
        try:
            book = load_workbook(file_path)
        except BadZipFile:
            book = Workbook()
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')
        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False):
            sheet.append(row)
        book.save(file_path)
    else:
        df.to_excel(file_path, index=False, sheet_name='Results', header=True)


log_file = open('run.log', 'a', encoding="utf-8")
def log(*args, **kwargs):
    print(*args, **kwargs)
    print(*args, file=log_file, **kwargs)
    log_file.flush()


class VarPool:
    def __init__(self):
        self._next = 1
        self._map = {}

    def new(self, key=None):
        if key is None:
            v = self._next
            self._next += 1
            return v
        if key in self._map:
            return self._map[key]
        v = self._next
        self._next += 1
        self._map[key] = v
        return v

# -------------------------- AMO_PE ------------------------------

def amo_product(cnf: CNF, pool: VarPool, lits):
    n = len(lits)
    if n <= 1:
        return
    p = int(math.ceil(math.sqrt(n)))
    q = int(math.ceil(n / p))
    # pad lits 
    L = lits + [None] * (p * q - n)
    rows = [[L[r * q + c] for c in range(q) if L[r * q + c] is not None] for r in range(p)]
    cols = [[L[r * q + c] for r in range(p) if L[r * q + c] is not None] for c in range(q)]

    # row selectors R_r, column selectors C_c
    R = [pool.new(("R", id(lits), r)) for r in range(p)]
    C = [pool.new(("C", id(lits), c)) for c in range(q)]

    # cell -> row/col selectors; and AMO over selectors (product shape)
    for r in range(p):
        for c in range(q):
            idx = r * q + c
            if idx >= len(L) or L[idx] is None:
                continue
            x = L[idx]
            # x -> R_r  and  x -> C_c
            cnf.append([-x, R[r]])
            cnf.append([-x, C[c]])

    # at most one row selected; at most one col selected
    def AMO_pairwise(vars_):
        k = len(vars_)
        for i in range(k):
            for j in range(i + 1, k):
                cnf.append([-vars_[i], -vars_[j]])

    AMO_pairwise(R)
    AMO_pairwise(C)


# ---------------------- Disjoint Intervals Encoder (Thm 22) ------------------
class DisjointIntervalsEncoder:
    def __init__(self, cnf: CNF, pool: VarPool):
        self.cnf = cnf
        self.pool = pool

    def _direct_encoding(self, X_ij: dict):
        # pairwise conflict for any overlapping intervals
        items = list(X_ij.items())
        L = len(items)
        for a in range(L):
            (i1, j1), v1 = items[a]
            for b in range(a + 1, L):
                (i2, j2), v2 = items[b]
                # intervals [i1,j1) and [i2,j2) overlap?
                if not (j1 <= i2 or j2 <= i1):
                    self.cnf.append([-v1, -v2])

    def encode(self, X_ij: dict, n: int):
        # X_ij: dict mapping (i,j) -> var id representing interval [i,j)
        if not X_ij:
            return
        # base-case threshold
        if n <= 20:
            self._direct_encoding(X_ij)
            return

        # choose number of blocks k ~ log n, block size b ~ n/k
        k = max(2, int(math.ceil(math.log2(n))))
        b = int(math.ceil(n / k))
        def B(x): return x // b

        # partition by block pairs
        by_blockpair = defaultdict(list)
        for (i, j) in X_ij.keys():
            bi, bj = B(i), B(j)
            by_blockpair[(bi, bj)].append((i, j))

        # recursive calls for each block-pair (remap indices)
        for (bl, br), pairs in by_blockpair.items():
            if not pairs:
                continue
            # collect the local domain (left block plus right block if distinct)
            # Always include all indices that appear in pairs
            endpoints = set()
            for (i, j) in pairs:
                endpoints.add(i)
                endpoints.add(j)

            if bl == br:
                block = range(bl * b, min((bl + 1) * b, n))
                local = sorted(set(block) | endpoints)
            else:
                left = range(bl * b, min((bl + 1) * b, n))
                right = range(br * b, min((br + 1) * b, n))
                local = sorted(set(left) | set(right) | endpoints)
            remap = {orig: idx for idx, orig in enumerate(local)}
            X_local = {(remap[i], remap[j]): X_ij[(i, j)] for (i, j) in pairs}
            self.encode(X_local, len(local))

        # group into Y block vars
        Y = {}
        group_pairs = defaultdict(list)
        for (i, j), v in X_ij.items():
            bi, bj = B(i), B(j)
            key = (bi, bj)
            group_pairs[key].append(v)
        # create Y vars and link
        for bl in range(k):
            for br in range(bl, k):
                key = (bl, br)
                Y[key] = self.pool.new(("Y", bl, br))
                xs = group_pairs.get(key, [])
                # X -> Y
                for x in xs:
                    self.cnf.append([-x, Y[key]])
                # Y -> OR X (if there are any)
                if xs:
                    self.cnf.append([-Y[key]] + xs)

        # enforce disjointness of Y as a smaller instance
        X_blk = { (bl, br): Y[(bl, br)] for bl in range(k) for br in range(bl, k) }
        # Recursively enforce on block-level
        self.encode(X_blk, k)

        
        # build list of actual intervals per block-pair to check overlaps
        block_to_intervals = defaultdict(list)
        for (i, j), v in X_ij.items():
            bi, bj = B(i), B(j)
            block_to_intervals[(bi, bj)].append((i, j, v))

        # For each pair of block-pairs, if any interval pair overlaps, we already had Y block conflicts.
        # So skip extra work here.

        return

class SchedulingEncoder:
    def __init__(self, N, M, T, d, r, e):
        self.N, self.M, self.T = N, M, T
        self.d, self.r, self.e = d, r, e
        self.pool = VarPool()
        self.cnf = CNF()
        self.x = {}  # (i,t,m)->var
        self.y = {}  # (m,t1,t2)->var

    # (C1)+(C2): create x and link to y
    def build_x_and_link(self):
        y_to_x = defaultdict(list)
        for i in range(self.N):
            t_min = max(0, self.r[i])
            t_max = min(self.T - self.d[i], self.e[i] - self.d[i])
            if t_min > t_max:
                # no feasible start -> UNSAT
                self.cnf.append([])  # empty clause
                return
            for t in range(t_min, t_max + 1):
                for m in range(self.M):
                    xv = self.pool.new(("x", i, t, m))
                    self.x[(i, t, m)] = xv
                    t1, t2 = t, t + self.d[i]
                    yk = (m, t1, t2)
                    yv = self.y.get(yk)
                    if yv is None:
                        yv = self.pool.new(("y", m, t1, t2))
                        self.y[yk] = yv
                    # Link (x -> y)
                    self.cnf.append([-xv, yv])
            
        # Add reverse link: y -> OR(x)   (Equation (4)-style)
        for yk, yv in self.y.items():
            xs = y_to_x.get(yk, [])
            if xs:
                self.cnf.append([-yv] + xs)

    # (C3): exactly-one start for each task (ALO + AMO_PE)
    def per_task_exactly_one(self):
        for i in range(self.N):
            lits = [v for (ii, t, m), v in self.x.items() if ii == i]
            if not lits:
                self.cnf.append([])  # UNSAT
                return
            # ALO
            self.cnf.append(lits)
            # AMO (product)
            amo_product(self.cnf, self.pool, lits)

    # (C4): AMO_PE per (m,t,Δ) bucket to forbid same-duration overlaps on (m,t)
    def duration_partition_amo(self):
        for m in range(self.M):
            for t in range(self.T):
                # group by duration
                groups = defaultdict(list)
                for i in range(self.N):
                    v = self.x.get((i, t, m))
                    if v is not None:
                        groups[self.d[i]].append(v)
                for Δ, lits in groups.items():
                    if len(lits) > 1:
                        amo_product(self.cnf, self.pool, lits)

    # (C5): disjointness of y-intervals per machine (Thm 22)
    def disjoint_y_intervals(self):
        # build intervals per machine
        intervals_by_machine = defaultdict(list)
        for (m, t1, t2), v in self.y.items():
            # sanity
            if 0 <= t1 < t2 <= self.T:
                intervals_by_machine[m].append((t1, t2, v))

        encoder = DisjointIntervalsEncoder(self.cnf, self.pool)
        for m, intervals in intervals_by_machine.items():
            # build X_ij mapping for this machine with indices 0..T (timepoints)
            X = {}
            for (t1, t2, v) in intervals:
                X[(t1, t2)] = v
            if X:
                encoder.encode(X, self.T)

    # Build final CNF
    def encode(self):
        self.build_x_and_link()
        # early UNSAT shortcut
        if any(len(c) == 0 for c in self.cnf.clauses):
            return self.cnf
        self.per_task_exactly_one()
        if any(len(c) == 0 for c in self.cnf.clauses):
            return self.cnf
        self.duration_partition_amo()
        # Disjoint intervals are required for correctness (Theorem 22)
        self.disjoint_y_intervals()
        return self.cnf

# ----------------------------- I/O and Solve ---------------------------------

def solve_from_instance(N, tasks, M=None):
    """
    tasks: list of (r_i, d_i, e_i)
    M: number of machines (default: len(tasks))
    """
    r = [t[0] for t in tasks]
    d = [t[1] for t in tasks]
    e = [t[2] for t in tasks]
    T = max(e) if e else 0
    M = M if M is not None else len(tasks)

    enc = SchedulingEncoder(N, M, T, d, r, e)
    start = time.time()
    cnf = enc.encode()
    # quick empty-clause UNSAT
    if any(len(c) == 0 for c in cnf.clauses):
        return False, "UNSAT by construction (some task infeasible)", None, None

    solver = Glucose3()
    for clause in cnf.clauses:
        solver.add_clause(clause)
    sat = solver.solve()
    elapsed = time.time() - start
    if not sat:
        return False, "UNSAT (solver)", None, elapsed, solver.nof_vars(), solver.nof_clauses()
    model = set(solver.get_model())

    # extract schedule from x
    schedule = []
    for (i, t, m), xv in enc.x.items():
        if xv in model:
            schedule.append((i, m, t, t + d[i]))
    schedule.sort()
    return True, schedule, enc, elapsed, solver.nof_vars(), solver.nof_clauses()

def solve_and_record_from_file(path, M_override=None):
    with open(path) as f:
        N = int(f.readline().strip())
        tasks = ast.literal_eval(f.readline().strip())
    ok, schedule_or_msg, enc, elapsed, nof_vars, nof_clauses = solve_from_instance(N, tasks, M_override)
    if not ok:
        return False, schedule_or_msg, elapsed, nof_vars, nof_clauses
    return True, schedule_or_msg, elapsed, nof_vars, nof_clauses

def process_input_dir(input_dir, folder):
    id = 1
    for filename in os.listdir(input_dir):
        if not filename.endswith(".txt"):
            continue
        path = os.path.join(input_dir, filename)
        result_container = {}
        finished_event = Event()
        time_budget = 120

        def solve_with_timeout(path, result_container, finished_event):
            try:
                ok, schedule_or_msg, elapsed, nof_vars, nof_clauses = solve_and_record_from_file(path)
                result_container['ok'] = ok
                result_container['schedule_or_msg'] = schedule_or_msg
                result_container['elapsed'] = elapsed
                result_container['nof_vars'] = nof_vars
                result_container['nof_clauses'] = nof_clauses
            except Exception as ex:
                result_container['error'] = ex
                result_container['traceback'] = traceback.format_exc()
            finally:
                finished_event.set()

        log(f"=== Processing {filename} ===")
        start_time = time.time()
        solver_thread = Thread(target=solve_with_timeout, args=(path, result_container, finished_event))
        solver_thread.start()

        finished = finished_event.wait(timeout=time_budget)
        solve_time = time.time() - start_time

        if not finished:
            log(f"✗ TIMEOUT after {time_budget}s")
            result_dict = {
                "ID": id,
                "Problem": filename,
                "Type": "biclique",
                "Time": round(solve_time, 4),
                "Result": "TIMEOUT",
                "Variables": nof_vars,
                "Clauses": nof_clauses
            }
            write_to_excel(result_dict, folder)
            solver_thread.join()
        elif 'error' in result_container:
            log("Error processing", filename, ":", result_container['error'])
            log(result_container.get('traceback', ''))
            result_dict = {
                "ID": id,
                "Problem": filename,
                "Type": "biclique",
                "Time": round(solve_time, 4),
                "Result": "ERROR",
                "Variables": nof_vars,
                "Clauses": nof_clauses
            }
            write_to_excel(result_dict, folder)
        else:
            ok = result_container['ok']
            schedule_or_msg = result_container['schedule_or_msg']
            elapsed = result_container['elapsed']
            nof_vars = result_container['nof_vars']
            nof_clauses = result_container['nof_clauses']
            if ok:
                log("SAT in %.3fs; schedule:" % (elapsed if elapsed else solve_time))
                for (i, m, s, f) in schedule_or_msg:
                    log(f" Task {i} -> machine {m}: [{s},{f})")
                    result_dict = {
                        "ID": id,
                        "Problem": filename,
                        "Type": "biclique",
                        "Time": round(elapsed if elapsed else solve_time, 4),
                        "Result": "SAT",
                        "Variables": nof_vars,
                        "Clauses": nof_clauses
                    }
            else:
                log("✗ NO FEASIBLE SCHEDULE EXISTS or error:", schedule_or_msg)
                result_dict = {
                    "ID": id,
                    "Problem": filename,
                    "Type": "biclique",
                    "Time": round(elapsed if elapsed else solve_time, 4),
                    "Result": "UNSAT",
                    "Variables": nof_vars,
                    "Clauses": nof_clauses
                }
            write_to_excel(result_dict, folder)

        id += 1
        
    log_file.close()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        log("Usage: script.py <input_subdir_under_input/>")
        sys.exit(1)
    sub = sys.argv[1]
    input_dir = os.path.join("input", sub)
    process_input_dir(input_dir, sub)
