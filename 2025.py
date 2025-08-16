import ast
import os
import time
import sys
import pandas as pd
import math
from collections import defaultdict
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from zipfile import BadZipFile
from pysat.solvers import Glucose3
from pysat.formula import CNF


log_file = open('run.log', 'a')
def log(*args, **kwargs):
    print(*args, **kwargs)
    # print(*args, file=log_file, **kwargs)
    log_file.flush()

def write_to_excel(result_dict):
    df = pd.DataFrame([result_dict])
    date = datetime.now().strftime('%Y-%m-%d')
    output_dir = 'out'
    if not os.path.exists(output_dir): os.makedirs(output_dir)
    file_path = f'{output_dir}/results_{date}_2025.xlsx'

    if os.path.exists(file_path):
        try:
            book = load_workbook(file_path)
        except BadZipFile:
            book = Workbook()
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')
        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False):
            sheet.append(row)
        book.save(file_path)
    else:
        df.to_excel(file_path, index=False, sheet_name='Results', header=True)

    log(f"Ghi kết quả vào Excel: {os.path.abspath(file_path)}")

def var_x(i, t, m, N, M, T):
    # times 0..T (use side = T+1)
    # index block per job: (T+1)*M
    return 1 + i * (M * (T + 1)) + t * M + m

def y_offset(N, M, T):
    return 1 + N * (M * (T + 1))

def var_y(m, t1, t2, N, M, T):
    # pack (t1,t2) with side = T+1
    side = T + 1
    return y_offset(N, M, T) + m * (side * side) + (t1 * side + t2)

def encode_schedule_problem(N, M, T, d, r, e):
    """
    Fixed encoding:
      - active_y is a dict mapping (m,t1,t2) -> yi to avoid duplicate y vars
      - only create y when at least one x maps to it (pruning)
      - use sequential counters (assumes add_at_most_one defined)
      - block decomposition: only add block-pair conflict if any interval-pair overlap
    """
    cnf = CNF()

    # quick exit
    if T <= 0:
        return cnf

    # helper: map for active y-variables (unique)
    active_y = {}              # (m,t1,t2) -> yi
    bucket = defaultdict(list) # (m,t1,t2) -> list of xi that imply this y
    x_lits_per_job = [[] for _ in range(N)]

    # 1) create x->y clauses and collect x-literals; create y only when needed
    for i in range(N):
        dur = d[i]
        t_start = max(0, r[i])
        t_end_allowed = min(e[i] - dur, T - dur)
        if t_start > t_end_allowed:
            # no feasible start for job i -> instance infeasible
            cnf.append([])  # empty clause => UNSAT
            return cnf
        for t1 in range(t_start, t_end_allowed + 1):
            t2 = t1 + dur
            for m in range(M):
                xi = var_x(i, t1, m, N, M, T)
                # ensure single y var per (m,t1,t2)
                if (m, t1, t2) not in active_y:
                    active_y[(m, t1, t2)] = var_y(m, t1, t2, N, M, T)
                yi = active_y[(m, t1, t2)]
                # x -> y
                cnf.append([-xi, yi])
                x_lits_per_job[i].append(xi)
                bucket[(m, t1, t2)].append(xi)

    # utility: sequential at-most-one (assumes cnf is pysat CNF object)
    def add_at_most_one(cnf_obj, lits):
        n = len(lits)
        if n <= 1:
            return
        s = [cnf_obj.nv + i + 1 for i in range(n-1)]
        cnf_obj.nv += n-1
        # x1 -> s1
        cnf_obj.append([-lits[0], s[0]])
        for idx in range(1, n-1):
            cnf_obj.append([-lits[idx], s[idx]])      # xi+1 -> si
            cnf_obj.append([-s[idx-1], s[idx]])       # si-1 -> si
            cnf_obj.append([-lits[idx], -s[idx-1]])   # xi+1 & si-1 -> false
        cnf_obj.append([-lits[n-1], -s[n-2]])        # xn & s_{n-1} -> false

    # 2) per-job exactly-one
    for i, lits in enumerate(x_lits_per_job):
        # at-least-one
        cnf.append(lits)
        # at-most-one via sequential
        add_at_most_one(cnf, lits)

    # 3) per-slot AMO: a y interval can be triggered by <=1 x
    for key, lits in bucket.items():
        add_at_most_one(cnf, lits)

    # Prepare intervals per machine for block decomposition
    intervals_by_machine = {m: [] for m in range(M)}
    # active_y items: (m,t1,t2) -> yi
    for (m, t1, t2), yi in active_y.items():
        # sanity check
        if 0 <= t1 < t2 <= T:
            intervals_by_machine[m].append((t1, t2, yi))

    # ensure cnf.nv at least as large as largest var id we used for x/y
    # estimate max x var id (safe upper bound)
    if N > 0:
        max_x_id = var_x(N-1, T, M-1 if M>0 else 0, N, M, T)
    else:
        max_x_id = 0
    max_y_id = max(active_y.values()) if active_y else 0
    cnf.nv = max(cnf.nv, max(max_x_id, max_y_id))

    # 4) block decomposition for non-overlap (heuristic)
    b = max(1, int(math.sqrt(max(1, T))))
    k = (T + b - 1) // b
    def block_of(t): return min(max(0, t // b), k-1)

    # group intervals into block-pairs per machine
    block_to_intervals = {}   # (m,l,r) -> list of (t1,t2,yi)
    for m in range(M):
        for (t1, t2, yi) in intervals_by_machine[m]:
            l = block_of(t1)
            rblk = block_of(t2 - 1)
            key = (m, l, rblk)
            block_to_intervals.setdefault(key, []).append((t1, t2, yi))

    # create yblock vars and link them to intervals in that block-pair
    yblock_id = {}  # (m,l,r) -> aux var id
    for key, intervals in block_to_intervals.items():
        # allocate aux var
        cnf.nv += 1
        yb = cnf.nv
        yblock_id[key] = yb
        # yi -> yb
        for (_, _, yi) in intervals:
            cnf.append([-yi, yb])
        # yb -> OR(yi's)
        clause = [-yb] + [yi for (_, _, yi) in intervals]
        cnf.append(clause)

    # 5) block-level conflicts: only add conflict if there exists an interval-pair that overlaps
    for m in range(M):
        # keys for this machine
        keys_m = [key for key in block_to_intervals.keys() if key[0] == m]
        for i_idx in range(len(keys_m)):
            key1 = keys_m[i_idx]
            yb1 = yblock_id[key1]
            ints1 = block_to_intervals.get(key1, [])
            for j_idx in range(i_idx + 1, len(keys_m)):
                key2 = keys_m[j_idx]
                yb2 = yblock_id[key2]
                ints2 = block_to_intervals.get(key2, [])
                # check any pair overlap
                overlap_exists = False
                for (a1, b1, yi) in ints1:
                    for (a2, b2, yj) in ints2:
                        if not (b1 <= a2 or b2 <= a1):
                            overlap_exists = True
                            break
                    if overlap_exists:
                        break
                if overlap_exists:
                    cnf.append([-yb1, -yb2])

    # 6) detailed intra-block conflicts (when l==r): explicit pairwise on real intervals inside the same block
    # Non-overlap: for each machine, forbid overlapping y-intervals
    for m in range(M):
        intervals = [ (t1,t2,yi) for (mm,t1,t2), yi in active_y.items() if mm == m ]
        for i in range(len(intervals)):
            t1, t2, yi = intervals[i]
            for j in range(i+1, len(intervals)):
                s1, s2, yj = intervals[j]
                if not (t2 <= s1 or s2 <= t1):  # overlap
                    cnf.append([-yi, -yj])


    return cnf


def solve_and_record(task_id, problem_name, N, M, T, d, r, e):
    log(f"Task({r}, {e}, {d}), Machines: {M}, Time slots: {T}")
    start = time.time()
    cnf = encode_schedule_problem(N, M, T, d, r, e)
    solver = Glucose3()
    for clause in cnf.clauses:
        solver.add_clause(clause)

    status = solver.solve()
    elapsed = time.time() - start

    if status:
        model = set(solver.get_model())
        for i in range(N):
            dur = d[i]
            t_start = max(0, r[i])
            t_end_allowed = min(e[i] - dur, T - dur)
            for t in range(t_start, t_end_allowed + 1):
                for m in range(M):
                    if var_x(i, t, m, N, M, T) in model:
                        log(f"Task {i} -> machine {m} : [{t}, {t+dur})")
                        found = True
                        # break
                        if found: 
                            break
    else:
        log("Không tìm được lịch hợp lệ.")

    result = {
        "ID": task_id,
        "Problem": problem_name,
        "Type": "biclique",
        "Time": round(elapsed, 4),
        "Result": "SAT" if status else "UNSAT",
        "Variables": solver.nof_vars(),
        "Clauses": solver.nof_clauses()
    }
    solver.delete()
    write_to_excel(result)

def process_input_dir(input_dir, resource=20):
    task_id = 1
    for filename in sorted(os.listdir(input_dir)):
        if filename.endswith(".txt"):
            path = os.path.join(input_dir, filename)
            with open(path) as f:
                N = int(f.readline().strip())
                tasks = ast.literal_eval(f.readline().strip())
                r = [task[0] for task in tasks]
                d = [task[1] for task in tasks]
                e = [task[2] for task in tasks]
                T = max(e)
                resource = len(tasks)
                log(f"\n Đang xử lý: {filename}")
                solve_and_record(task_id, filename, N, resource, T, d, r, e)
                task_id += 1
    log_file.close()

if __name__ == "__main__":
    input_dir = os.path.join("input", sys.argv[1])
    process_input_dir(input_dir)