import os
import ast
import sys
import time
import pandas as pd
from threading import Thread, Event
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pysat.solvers import Glucose3
from zipfile import BadZipFile
from datetime import datetime

OUTPUT_DIR = "out/"
LOG_FILE = open("run.log", "a", encoding="utf-8")
SOLVER_CLASS = Glucose3
TIME_LIMIT = 1200
ENCODING_TYPE = "es3"

def log(*args, **kwargs):
    print(*args, **kwargs)
    print(*args, file=LOG_FILE, **kwargs)
    LOG_FILE.flush()

def ensure_output_dir():
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)

def read_input_file(filepath: str):
    with open(filepath, "r") as f:
        try:
            num_tasks = int(f.readline().strip())
            tasks = ast.literal_eval(f.readline().strip())
            assert isinstance(tasks, list) and all(isinstance(t, tuple) and len(t) == 3 for t in tasks)
            return tasks
        except Exception as e:
            raise ValueError(f"Lỗi khi đọc file {filepath}: {e}")

def encode_es3(tasks, resources, solver):
    max_time = max(d for _, _, d in tasks)
    n_tasks = len(tasks)

    u = [[i * resources + j + 1 for j in range(resources)] for i in range(n_tasks)]
    z = [[n_tasks * resources + i * max_time + t + 1 for t in range(tasks[i][0], tasks[i][2])] for i in range(n_tasks)]
    D = [[[n_tasks * (resources + max_time) + i * resources * max_time + j * max_time + t + 1
           for t in range(tasks[i][0], tasks[i][2] - tasks[i][1] + 1)]
          for j in range(resources)] for i in range(n_tasks)]

    # D1: Add clauses to ensure each task uses exactly one resource
    for i in range(n_tasks):
        for j in range(resources):
            for jp in range(j + 1, resources):
                solver.add_clause([-u[i][j], -u[i][jp]])

    # D2: Add clauses to ensure each task is assigned to at least one resource
    for i in range(n_tasks):
        solver.add_clause([u[i][j] for j in range(resources)])

    # D3: Add clauses to ensure no two tasks overlap on the same resource
    for i in range(n_tasks):
        for ip in range(i + 1, n_tasks):
            for j in range(resources):
                for t in range(max(tasks[i][0], tasks[ip][0]), min(tasks[i][2], tasks[ip][2])):
                    solver.add_clause([
                        -z[i][t - tasks[i][0]], -u[i][j],
                        -z[ip][t - tasks[ip][0]], -u[ip][j]
                    ])

    # D4: Add clauses to ensure each task is executed at least once
    for i in range(n_tasks):
        clause = []
        for j in range(resources):
            for t in range(tasks[i][0], tasks[i][2] - tasks[i][1] + 1):
                clause.append(D[i][j][t - tasks[i][0]])
        solver.add_clause(clause)

    # D5: Add clauses to ensure each task is executed for its duration
    for i, (r, e, d) in enumerate(tasks):
        for j in range(resources):
            for t in range(r, d - e + 1):
                tidx = t - r
                solver.add_clause([-D[i][j][tidx], u[i][j]])
                for tp in range(t, t + e):
                    solver.add_clause([-D[i][j][tidx], z[i][tp - r]])
                for tp in range(r, t):
                    solver.add_clause([-D[i][j][tidx], -z[i][tp - r]])
                for tp in range(t + e, d):
                    solver.add_clause([-D[i][j][tidx], -z[i][tp - r]])

    return u, z, D

def validate(tasks, model, u, z, D, resources):
    task_resource = {}
    task_time = {}
    resource_usage = {j: set() for j in range(resources)}

    for i, (r, e, d) in enumerate(tasks):
        task_time[i] = [t + r for t in range(d - r) if model[z[i][t] - 1] > 0]
        for j in range(resources):
            if model[u[i][j] - 1] > 0:
                task_resource[i] = j

    for i in range(len(tasks)):
        if i not in task_resource:
            log(f"[INVALID] Task {i+1} không có tài nguyên.")
            return False
        times = task_time[i]
        if not times or len(times) != tasks[i][1]:
            log(f"[INVALID] Task {i+1} không thực thi đúng thời lượng.")
            return False
        if sorted(times) != list(range(times[0], times[0] + tasks[i][1])):
            log(f"[INVALID] Task {i+1} không thực thi liên tục.")
            return False
        for t in times:
            if t in resource_usage[task_resource[i]]:
                log(f"[INVALID] Tài nguyên bị trùng thời gian tại time {t}.")
                return False
            resource_usage[task_resource[i]].add(t)

    log("[VALID] Lịch thực thi hợp lệ.")
    return True

def solve_es3(tasks, resources):
    solver = SOLVER_CLASS()
    result = {}
    event = Event()

    def worker():
        try:
            u, z, D = encode_es3(tasks, resources, solver)
            status = solver.solve()
            if status:
                model = solver.get_model()
                if model and validate(tasks, model, u, z, D, resources):
                    result.update({'status': 'SAT', 'model': model, 'u': u, 'z': z, 'D': D})
                else:
                    result.update({'status': 'INVALID'})
            else:
                result.update({'status': 'UNSAT'})
        except Exception as e:
            result.update({'status': 'ERROR', 'error': str(e)})
        finally:
            event.set()

    thread = Thread(target=worker)
    start = time.time()
    thread.start()
    finished = event.wait(timeout=TIME_LIMIT)
    elapsed = time.time() - start

    if not finished:
        solver.interrupt()
        thread.join()
        return "TIMEOUT", elapsed, 0, 0

    status = result.get('status', 'ERROR')
    model = result.get('model')
    num_vars = solver.nof_vars()
    num_clauses = solver.nof_clauses()
    solver.delete()

    if status == "SAT":
        return "SAT", elapsed, num_vars, num_clauses
    else:
        return status, elapsed, num_vars, num_clauses

def write_result_to_excel(result_dict, input_folder):
    ensure_output_dir()
    df = pd.DataFrame([result_dict])
    today = datetime.now().strftime('%Y-%m-%d')
    input_folder_name = os.path.basename(input_folder.replace("input/", ""))
    excel_path = os.path.join(OUTPUT_DIR, f'results_{today}_{input_folder_name}.xlsx')

    if os.path.exists(excel_path):
        try:
            book = load_workbook(excel_path)
        except BadZipFile:
            book = Workbook()
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')
        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False):
            sheet.append(row)
        book.save(excel_path)
    else:
        df.to_excel(excel_path, index=False, sheet_name='Results', header=True)
    log(f"Đã ghi kết quả vào: {excel_path}")

def run_batch(input_folder, resources=50):
    case_id = 1
    for filename in sorted(os.listdir(input_folder)):
        if not filename.endswith(".txt"):
            continue
        filepath = os.path.join(input_folder, filename)
        log(f"\n=== Đang xử lý {filename} ===")
        try:
            tasks = read_input_file(filepath)
            log(f"Tác vụ: {tasks}")
            status, duration, vars_used, clauses = solve_es3(tasks, resources)
            result = {
                "ID": case_id,
                "Problem": filename,
                "Type": ENCODING_TYPE,
                "Time": round(duration, 3),
                "Result": status,
                "Variables": vars_used,
                "Clauses": clauses
            }
            write_result_to_excel(result, input_folder)
            case_id += 1
        except Exception as e:
            log(f"[LỖI] Không thể xử lý file {filename}: {e}")

if __name__ == "__main__":
    folder = "medium"
    run_batch("input/" + folder)
    LOG_FILE.close()
